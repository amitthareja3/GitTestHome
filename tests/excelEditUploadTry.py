import openpyxl
from selenium import webdriver
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

#.....Chrome
Service_obj = Service("C:\\webdrivers\\chromedriver.exe")
driver = webdriver.Chrome(service=Service_obj)
#
# Firfox
# Service_obj = Service("C:\\webdrivers\\geckodriver.exe")
# driver = webdriver.Firefox(service=Service_obj)
driver.maximize_window()
def update_excel_data(filepath,fruit_name,colname,new_value):
        Dict = {}

        book = openpyxl.load_workbook(filepath)
        sheet = book.active

        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=i).value == colname:
                Dict["col"] = i

        for j in range(1, sheet.max_row + 1):
            for k in range(1, sheet.max_column + 1):
                if sheet.cell(row=j, column=k).value == fruit_name:
                    Dict["row"] = j

        sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
        book.save(filepath)

# upload in excel
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(5)
newValue = 189
filepath = "C:\\download.xlsx"
fruit_name = "Apple"
#
# Edit method called here
update_excel_data(filepath, fruit_name,"price",newValue)

# driver.find_element(By.ID,"downloadButton").click()

# upload done here
file_input = driver.find_element(By.CSS_SELECTOR, "input[id=fileinput]")
file_input.send_keys(filepath)
toast_Locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait = WebDriverWait(driver,5)
wait.until(expected_conditions.visibility_of_element_located(toast_Locator))
print(driver.find_element(*toast_Locator).text)
priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[contains(text(),'"+fruit_name+"')]/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
# or xpath can be thus      //div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']

assert actual_price == newValue



# below is my try will check and run some other time

# class excelTry:
#     def getTestExcelData(self):
#         Dict = {}
#         # book = openpyxl.load_workbook(filepath)
#         book = openpyxl.load_workbook("C:\\download.xlsx")
#         # book = openpyxl.load_workbook("C:\\Users\\Amit Thareja\\OneDrive\\Documents\\pythonDemo.xlsx")
#         sheet = book.active
#         for i in range(1, sheet.max_column + 1):
#              if sheet.cell(row=1, column=i).value == "fruit_name":
#                  for j in range(1, sheet.max_row +1 ):
#                     if sheet.cell(row=j, column=i).value == "Apple":
#                         # print(sheet.cell(row=j, column=i).value)
#                         for k in range(2, sheet.max_column + 1):
#                             Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=j, column=k).value
#
#          print(Dict)
#
