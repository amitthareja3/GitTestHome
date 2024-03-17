import openpyxl
Dict={}
book = openpyxl.load_workbook("C:\\Users\\Amit Thareja\\OneDrive\\Documents\\pythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
# print(cell.value)


for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase2":  # if needed to print row where value starts with TestCase2 in sheet
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
            # print(sheet.cell(row=i,column=j).value)

print(Dict)
