import openpyxl


class HomePageData:
    #  option 1 - below was ealier way by which data was sent
    test_HomePageData = [{"firstname":"Rahul","lastname":"Shetty","gender":"Male"}, {"firstname":"Anushika","lastname":"Shetty","gender":"Female"},{"firstname":"Mohan","lastname":"Sharma","gender":"Male"}]
    # option 2 - below is the way data taken from excel
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Amit Thareja\\OneDrive\\Documents\\pythonDemo.xlsx")
        # book = openpyxl.load_workbook("C:\\pythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # if needed to print row where value starts with TestCase2 in sheet
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

